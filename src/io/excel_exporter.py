try:
	from openpyxl import Workbook
	from openpyxl.styles import Font
except ImportError:
	raise ImportError(
		"Brak biblioteki openpyxl. Zainstaluj: pip install openpyxl"
	)

class ExcelExporter:
    def export(self, dataset, statistics, filepath: str) -> None:
        try:
            wb = Workbook()
        except OSError as e:
            raise IOError(f"Blad zapisu pliku: {e}")

        self._create_transactions_sheet(wb, dataset)
        self._create_statistics_sheet(wb, dataset, statistics)
        self._create_products_sheet(wb, dataset)

        wb.save(filepath)

    def _create_transactions_sheet(self, wb, dataset):
        ws = wb.active

        headers = [
            "Data",
            "Produkt",
            "Kategoria",
            "Ilosc",
            "Cena jedn.",
            "Wartosc",
            "Sprzedawca",
            "Region",
        ]

        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        ws.freeze_panes = "A2"

        records = sorted(dataset, key=lambda r: r.date)

        for record in records:
            ws.append([
                record.date.strftime("%Y-%m-%d"),
                record.product.name,
                record.product.category,
                record.quantity,
                record.product.price,
                record.total_value(),
                record.seller,
                record.region
            ])

        for col in ["E", "F"]:
            for cell in ws[col]:
                if cell.row > 1:
                    cell.number_format = "#,##0.00"

        self._auto_width(ws)

    def _create_statistics_sheet(self, wb, dataset, statistics):
        ws = wb.create_sheet("Statystyki")
        
        def _write_title(ws, row, text):
            cell = ws.cell(row=row, column=1, value=text)
            cell.font = Font(bold=True)

            return row + 1

        def _write_kv(ws, row, label, value):
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)

            return row + 1

        def _write_section(ws, row, title, data):
            row = _write_title(ws, row, title)

            for item in data:
                k, v = item
                row = _write_kv(ws, row, k, v)

            return row + 1

        row = 1
        
        row = _write_title(ws, row, "PODSUMOWANIE")
        row = _write_kv(ws, row, "Łączny przychód", statistics.total_revenue())
        row = _write_kv(ws, row, "Liczba transakcji", len(dataset))
        row = _write_kv(ws, row, "Średnia transakcja", statistics.average_transaction())

        row += 1

        row = _write_section(ws, row, "PRZYCHÓD WG KATEGORII", statistics.by_category().items())
        row = _write_section(ws, row, "PRZYCHÓD WG SPRZEDAWCY", statistics.by_seller().items())
        row = _write_section(ws, row, "TOP 5 PRODUKTÓW", statistics.top_products(5))
        row = _write_section(ws, row, "PODSUMOWANIE MIESIĘCZNE", statistics.monthly_summary().items())


    def _create_products_sheet(self, wb, dataset):
        ws = wb.create_sheet("Produkty")

        headers = [
            "ID",
            "Nazwa",
            "Kategoria",
            "Cena (PLN)"
        ]

        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        products = {}

        for record in dataset:
            product = record.product
            products[product.product_id] = product

        for product in sorted(products.values(), key=lambda p: p.product_id):
            ws.append([
                product.product_id,
                product.name,
                product.category,
                product.price
            ])

        for cell in ws["D"]:
            if cell.row > 1:
                cell.number_format = "#,##0.00"

        self._auto_width(ws)

    def _auto_width(self, ws):
        for column in ws.columns:
            column_letter = column[0].column_letter

            max_length = 0

            for cell in column:
                value = cell.value
                if value is None:
                    continue

                max_length = max(max_length, len(str(value)))

            ws.column_dimensions[column_letter].width = max_length * 2